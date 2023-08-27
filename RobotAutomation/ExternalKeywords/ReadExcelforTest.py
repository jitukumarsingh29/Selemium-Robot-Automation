import openpyxl

wk = openpyxl.load_workbook("C:/VirtualDisk-D/Study/Automation Testing/Robot Framework/TC_005_DDT.xlsx")


def fetch_no_of_rows(sheetname):
    sh = wk[sheetname]
    return sh.max_row


def fetch_cell_data(sheetname, row, cell):
    sh = wk[sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value


print(fetch_no_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1", 1, 1))
