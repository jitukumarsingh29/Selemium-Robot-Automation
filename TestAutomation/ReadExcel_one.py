import openpyxl

wk = openpyxl.load_workbook("C:\\VirtualDisk-D\\Study\\TestingExcel.xlsx")
sh = wk['Sheet1']
print(sh['A1'].value)
print(sh['B1'].value)
print(sh['C1'].value)
print(sh['D1'].value)
print("********************************")
c1 = sh.cell(3,1)
print(c1.value)
print("********************************")
c2 = sh.cell(column=1, row=3)
print(c2.value)
print(c2.row)
print(c2.column)
print("********************************")
rows = sh.max_row
print(rows)
column = sh.max_column
print(sh.max_column)
print("********************************")
for i in range(1, rows+1):
    for j in range(1, column+1):
        c = sh.cell(i, j)
        print(c.value)
print("********************************")
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)
print("********************************")